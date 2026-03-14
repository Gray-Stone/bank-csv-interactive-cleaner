"""CSV and XLSX loading helpers for tabular bank exports."""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SUPPORTED_SUFFIXES = {".csv", ".xlsx"}


@dataclass
class LoadedSheet:
    """Represent a loaded tabular file before parsing."""
    file_path: Path
    file_type: str
    sheet_name: str
    rows: list[list[str]]
    encoding_guess: str = "utf-8-sig"
    delimiter_guess: str = ","


def _normalize_cell(value: Any) -> str:
    """Internal helper for normalize cell."""
    if value is None:
        return ""
    return str(value).strip()


def _detect_csv_encoding(raw_bytes: bytes) -> str:
    """Internal helper for detect csv encoding."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            raw_bytes.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detect_delimiter(sample: str) -> str:
    """Internal helper for detect delimiter."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def load_csv(path: Path) -> LoadedSheet:
    """Load csv."""
    raw_bytes = path.read_bytes()
    encoding = _detect_csv_encoding(raw_bytes)
    text = raw_bytes.decode(encoding, errors="replace")
    sample = "\n".join(text.splitlines()[:20])
    delimiter = _detect_delimiter(sample)
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    rows = [[_normalize_cell(cell) for cell in row] for row in reader]
    return LoadedSheet(
        file_path=path,
        file_type="csv",
        sheet_name=path.name,
        rows=rows,
        encoding_guess=encoding,
        delimiter_guess=delimiter,
    )


def load_xlsx(path: Path, sheet_name: str | None = None) -> LoadedSheet:
    """Load xlsx."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    target_sheet = sheet_name or workbook.sheetnames[0]
    worksheet = workbook[target_sheet]
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append([_normalize_cell(cell) for cell in row])
    return LoadedSheet(
        file_path=path,
        file_type="xlsx",
        sheet_name=target_sheet,
        rows=rows,
        encoding_guess="utf-8",
        delimiter_guess=",",
    )


def load_tabular_file(input_file: str, sheet_name: str | None = None) -> LoadedSheet:
    """Load tabular file."""
    path = Path(input_file).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"Unsupported file type '{suffix}'. Supported: {sorted(SUPPORTED_SUFFIXES)}")
    if suffix == ".csv":
        return load_csv(path)
    return load_xlsx(path, sheet_name=sheet_name)
